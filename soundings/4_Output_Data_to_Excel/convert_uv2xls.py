### NAME:  convert_uv2xls.py

### MODIFICATION HISTORY:  Written by Maiana Hanshaw for Python (03/28/2020);

### PURPOSE:  To calculate our own shear values from various datasets and put them into excel format.

### RESTRICTIONS:  Data has to be in the UV format as shown below:
#Project: 				VSE-2018
#Platform ID/Location: 	NWS: KBMX
#Date/Time (UTC): 		20180328/2306
#Latitude/Longitude: 	33.18000/-86.78300
#Altitude (masl): 		174.0
#---------------------------------------------
#HEIGHT(masl)  WSPD(m/s)  WDIR  U(m/s)  V(m/s)
#   174.0         5.1    159.4   -1.8     4.8

###############################################################################

import os  # operating system library
import pandas as pd  # pandas library for dictionary and data frames
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, NamedStyle, Side

###### UPDATE THIS ######
directory_in = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO/CSU/UV_Files"  # location of UV text files
directory_out = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO"  # location to output excel file
file_out = "RELAMPAGO_CSU_IOP04_UV.xlsx"
directory_in_problemfile = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO/CSU/UV_Files"  # location of sounding file problem list
problem_file_name = "Problem_Files.txt"
invalid_value = -9999
#########################

def get_iop(date):

    iop = ""
    # PECAN
    if int(date) == 20150624 or int(date) == 20150625:
        iop = "15"
    if int(date) == 20150705 or int(date) == 20150706:
        iop = "20"
    if int(date) == 20150712 or int(date) == 20150713:
        iop = "27"

    # VSE-2017
    if int(date) == 20170327 or int(date) == 20170328:
        iop = "1B"
    if int(date) == 20170405 or int(date) == 20170406:
        iop = "3B"
    if int(date) == 20170430 or int(date) == 20170501:
        iop = "4C"
        
    # VSE-2018
    if int(date) == 20180328 or int(date) == 20180329:
        iop = "2A"
    if int(date) == 20180403 or int(date) == 20180404:
        iop = "3"
    if int(date) == 20180406 or int(date) == 20180407:
        iop = "4"

    # RELAMPAGO        
    if int(date) == 20181102:
        iop = "1"
    if int(date) == 20181105:
        iop = "2"
    if int(date) == 20181106:
        iop = "3"        
    if int(date) == 20181110:
        iop = "4"           
    if int(date) == 20181111 or int(date) == 20181112:
        iop = "5"        
    if int(date) == 20181117:
        iop = "6"    
    if int(date) == 20181121:
        iop = "7"   
    if int(date) == 20181122:
        iop = "8"   
    if int(date) == 20181125:
        iop = "9"   
    if int(date) == 20181126:
        iop = "10"
    if int(date) == 20181129:
        iop = "11"
    if int(date) == 20181130:
        iop = "12"
    if int(date) == 20181204:
        iop = "13"
    if int(date) == 20181205:
        iop = "14"
    if int(date) == 20181210:
        iop = "15"
    if int(date) == 20181211:
        iop = "16"
    if int(date) == 20181213 or int(date) == 20181214:
        iop = "17"
    if int(date) == 20181215:
        iop = "18"
    if int(date) == 20181216:
        iop = "19"

    return iop

#########################

def get_files_from_directory(directory_in):

    selected_files = []
    for root, dirs, files in os.walk(directory_in):
        if root == directory_in:
            for file in files:
                if "UV" in file:
                    selected_files += [file]
    return selected_files

#########################
    
def open_file_and_split_into_lines(file_in):
    file_lines = []
    with open (os.path.join(directory_in, file_in), "r") as myfile:
        for file_line in myfile:          
            file_lines.append(file_line)
    return file_lines

#########################
    
def get_list_of_problem_files(directory_in, problem_file_name):

    file_lines = []    
    if os.path.exists(os.path.join(directory_in, problem_file_name)):
        with open (os.path.join(directory_in, problem_file_name), "r") as myfile:
            for file_line in myfile:          
                file_lines.append(file_line)
    return file_lines

#########################

def uv_at_h0(height_values, u_values, v_values):        

    # If -9999 at the lowest height, look to the next height
    non_invalid_inds = [ind for ind in range(len(height_values)) if (u_values[ind] != invalid_value and v_values[ind] != invalid_value)]    
    first_valid_ind = non_invalid_inds[0]
    
    u0 = u_values[first_valid_ind]
    v0 = v_values[first_valid_ind]
    h0 = height_values[first_valid_ind]
    
    return u0, v0, h0

#########################

def uv_at_closest_height(height_values, altitude_level, u_values, v_values):

    if altitude_level > height_values[-1]:
        height_closest = ""
        height_index = ""
        u_height_closest = ""
        v_height_closest = ""
        
    else:   
        height_closest = height_values[min(range(len(height_values)), key=lambda i: abs(height_values[i]-altitude_level))]
        height_index = height_values.index(height_closest)
        u_height_closest = u_values[height_index]
        v_height_closest = v_values[height_index]
        
    if u_height_closest == -9999 or v_height_closest == -9999:
        height_closest_range_min = height_values[min(range(len(height_values)), key=lambda i: abs(height_values[i]-(altitude_level - 50)))]
        height_closest_range_max = height_values[min(range(len(height_values)), key=lambda i: abs(height_values[i]-(altitude_level + 50)))]
        height_index_range_min = height_values.index(height_closest_range_min)
        height_index_range_max = height_values.index(height_closest_range_max)
            
        h_values_range = height_values[height_index_range_min: height_index_range_max]
        u_values_range = u_values[height_index_range_min: height_index_range_max]
        v_values_range = v_values[height_index_range_min: height_index_range_max]

        height_closest_range = []        
        for ind in range(len(h_values_range)):
            if u_values_range[ind] == -9999 or v_values_range[ind] == -9999:
                continue
            height_closest_range.append(h_values_range[ind])
        if height_closest_range == []:
            height_closest = ""
            height_index = ""
            u_height_closest = ""
            v_height_closest = ""
        else:
            height_closest = height_closest_range[min(range(len(height_closest_range)), key=lambda i: abs(height_closest_range[i]-altitude_level))]
            height_index = height_values.index(height_closest)
            u_height_closest = u_values[height_index]
            v_height_closest = v_values[height_index]
      
    return height_closest, height_index, u_height_closest, v_height_closest
    
#########################

def parse_info_from_uv_file(file_in):

    print(file_in)  

    # Extract site information
    file_lines = open_file_and_split_into_lines(file_in)
    
    project = file_lines[0].split(":", 1)[1].strip()
    name = file_lines[1].split(":", 1)[1].strip()
    date = file_lines[2].split(":", 1)[1].strip().split("/")[0]
    t = file_lines[2].split(":", 1)[1].strip().split("/")[1]
    time = t[0:2] + ":" + t[2:4]
    lat = file_lines[3].split(":", 1)[1].strip().split("/")[0]
    lon = file_lines[3].split(":", 1)[1].strip().split("/")[1]
    altitude = float(file_lines[4].split(":", 1)[1].strip())
        
    # Get IOP #
    iop = get_iop(date)   
        
    #########################

    # Extract data header and data into a dictionary and then a data frame
    data_d = pd.read_csv(os.path.join(directory_in, file_in), sep="\s{1,}", engine="python", header=6, usecols=["HEIGHT(masl)", "U(m/s)", "V(m/s)"]).to_dict(orient="list")
    data_df = pd.DataFrame.from_dict(data_d, orient='columns').astype(float).sort_index()  # convert dictionary to a data frame with float numbers

    # Get Height in MAGL not MASL    
    data_df["HEIGHT(masl)"] -= altitude  # remove initial altitude to height values to get MAGL
    data_df.columns = ["HEIGHT(magl)", "U(m/s)", "V(m/s)"]

    h_list = data_df["HEIGHT(magl)"].values.tolist()
    u_list = data_df["U(m/s)"].values.tolist()
    v_list = data_df["V(m/s)"].values.tolist()

    # Get u and v values at different heights
    u0, v0, h0 = uv_at_h0(h_list, u_list, v_list)

    h1, h1_ind, u1, v1 = uv_at_closest_height(h_list, 1000, u_list, v_list)
    h3, h3_ind, u3, v3 = uv_at_closest_height(h_list, 3000, u_list, v_list)
    h6, h6_ind, u6, v6 = uv_at_closest_height(h_list, 6000, u_list, v_list)
    h8, h8_ind, u8, v8 = uv_at_closest_height(h_list, 8000, u_list, v_list)
    h9, h9_ind, u9, v9 = uv_at_closest_height(h_list, 9000, u_list, v_list)
      
    #########################
  
    file_lines = get_list_of_problem_files(directory_in_problemfile, problem_file_name)
    problem = ""
    for ind in range(len(file_lines)):
        if name in file_lines[ind] and date in file_lines[ind] and t in file_lines[ind]:
            problem += file_lines[ind].split(":", 1)[1]
            
    #If first height is > 15m, note it as a problem
    if h0 > 15:
        problem += "PROBLEM: U0, V0 from height (magl) = " + '{:.1f}'.format(h0)

    #########################    
        
    # Get data in the right format for writing to excel:
    header_values = ["IOP", "Date", "Time", "Site", "Lat", "Lon", "U0", "V0", "U1", "V1", "U3", "V3", "V6", "V6", "U8", "V8", "U9", "V9", "Sounding Issues"]
    data_values = [iop, date, time, name, lat, lon, u0, v0, u1, v1, u3, v3, u6, v6, u8, v8, u9, v9, problem]

    return header_values, data_values

#########################

def write_data_to_xlsx_worksheet(workbook, worksheet, data_values, row_number):

    c = 1
    for data in data_values:
        worksheet.cell(row=row_number, column=c).value = data
        c += 1
        
    return workbook
    
#########################

def format_xlsx_worksheet(workbook, worksheet, header_values, number_of_rows):

    # Header
    header = NamedStyle(name="header")
    header.font = Font(name="Calibri", size=14, bold=True)
    header.border = Border(bottom=Side(border_style="medium"))
    header.alignment = Alignment(horizontal="center")

    c = 1
    for header_name in header_values:
        worksheet.cell(row=2, column=c).value = header_name
        c += 1
     
    for cell in worksheet["2:2"]:
        cell.style = header
    
    # Data
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name="Calibri", size=14, bold=False)
    data_style.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(3,number_of_rows):
        for cell in row:
            cell.style = data_style
        
    return workbook

###############################################################################

files_to_process = get_files_from_directory(directory_in)
number_of_rows = len(files_to_process) + 2
wb = Workbook()
ws = wb.active
row_number = 3

for file in files_to_process:
    header_values, data_values = parse_info_from_uv_file(file)
    write_data_to_xlsx_worksheet(wb, ws, data_values, row_number)
    row_number += 1

format_xlsx_worksheet(wb, ws, header_values, number_of_rows)
wb.save(os.path.join(directory_out, file_out))

###############################################################################