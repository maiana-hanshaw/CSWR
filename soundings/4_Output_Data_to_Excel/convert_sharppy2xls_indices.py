### NAME:  convert_sharppy2xls_indices.py

### MODIFICATION HISTORY:  Written by Maiana Hanshaw for Python (03/27/2020);

### PURPOSE:  To get the SHARPpy output indices and put them into excel format.

###############################################################################

import os  # operating system library
import re  # regular expressions library
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, NamedStyle, Side

###### UPDATE THIS ######
directory_in = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO/CSU/SHARPpy/Indices"  # location of SHARPpy Indices text files
directory_out = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO"  # location to output excel file
file_out = "RELAMPAGO_CSU_IOP04_Indices.xlsx"
directory_in_problemfile = "C:/Users/Maiana/Downloads/Soundings/RELAMPAGO/CSU/SPC_Files"  # location of sounding file problem list
problem_file_name = "Problem_Files.txt"
#########################

def get_iop(date):

    iop = ""
    # PECAN
    if int(date) == 20150624 or int(date) == 20150625:
        iop = "15"
    if int(date) == 20150705 or int(date) == 20150706:
        iop = "20"
    if int(date) == 20150712 or int(date) == 20150713:
        iop = "27"

    # VSE-2017
    if int(date) == 20170327 or int(date) == 20170328:
        iop = "1B"
    if int(date) == 20170405 or int(date) == 20170406:
        iop = "3B"
    if int(date) == 20170430 or int(date) == 20170501:
        iop = "4C"
        
    # VSE-2018
    if int(date) == 20180328 or int(date) == 20180329:
        iop = "2A"
    if int(date) == 20180403 or int(date) == 20180404:
        iop = "3"
    if int(date) == 20180406 or int(date) == 20180407:
        iop = "4"
        
    # RELAMPAGO        
    if int(date) == 20181102:
        iop = "1"
    if int(date) == 20181105:
        iop = "2"
    if int(date) == 20181106:
        iop = "3"        
    if int(date) == 20181110:
        iop = "4"           
    if int(date) == 20181111 or int(date) == 20181112:
        iop = "5"        
    if int(date) == 20181117:
        iop = "6"    
    if int(date) == 20181121:
        iop = "7"   
    if int(date) == 20181122:
        iop = "8"   
    if int(date) == 20181125:
        iop = "9"   
    if int(date) == 20181126:
        iop = "10"
    if int(date) == 20181129:
        iop = "11"
    if int(date) == 20181130:
        iop = "12"
    if int(date) == 20181204:
        iop = "13"
    if int(date) == 20181205:
        iop = "14"
    if int(date) == 20181210:
        iop = "15"
    if int(date) == 20181211:
        iop = "16"
    if int(date) == 20181213 or int(date) == 20181214:
        iop = "17"
    if int(date) == 20181215:
        iop = "18"
    if int(date) == 20181216:
        iop = "19"

    return iop

#########################

def get_files_from_directory(directory_in):

    selected_files = []
    for root, dirs, files in os.walk(directory_in):
        if root == directory_in:
            for file in files:
                if "Indices" in file:
                    selected_files += [file]
    return selected_files

#########################
    
def get_list_of_problem_files(directory_in, problem_file_name):

    file_lines = []    
    if os.path.exists(os.path.join(directory_in, problem_file_name)):
        with open (os.path.join(directory_in, problem_file_name), "r") as myfile:
            for file_line in myfile:          
                file_lines.append(file_line)
    return file_lines
    
#########################

def parse_info_from_indices_file(file_in):

    print(file_in)  

    # Extract site info from file name
    site_info = re.split(r'[_.\s]\s*', file_in)
 
    # Get site id/name
    if "RELAMPAGO" in directory_in:
        name = site_info[2]
        inst_id = name
        
    else:        
        inst_id = site_info[1]
        if len(site_info) == 5:
            name = site_info[1]
        if len(site_info) == 6:
            name = site_info[1] + ": " + site_info[2]
        if len(site_info) == 7:
            name = site_info[1] + ": " + site_info[2] + " " +site_info[3]
        
    # Get date and time
    if "RELAMPAGO" in directory_in:
        date = site_info[1]
        t = site_info[3]
        time = t[0:2] + ":" + t[2:4]

    else:        
        if len(site_info) == 5:
            date = site_info[2]
            t = site_info[3]
            time = t[0:2] + ":" + t[2:4]
        if len(site_info) == 6:
            date = site_info[3]
            t = site_info[4]
            time = t[0:2] + ":" + t[2:4]
        if len(site_info) == 7:
            date = site_info[4]
            t = site_info[5]
            time = t[0:2] + ":" + t[2:4]
              
    # Get IOP #
    iop = get_iop(date)

    #########################

    # Convert indices text file to dictionary:   
    with open(os.path.join(directory_in, file_in),'r') as f:
        indices_dict = eval(f.read())
        
    # Get data into new variables and convert to different units if necessary (eg. knots to m/s for wind speed):
    lat = ""
    lon = ""
    srh_1km = ""
    srh_3km = ""
    srh_1km_abs = ""
    srh_3km_abs = ""
    cape_sb = ""
    cape_ml = ""
    cape_mu = ""
    lcl_sb = ""
    lcl_ml = ""
    shear_1km = ""
    shear_3km = ""
    shear_6km = ""
    shear_8km = ""
    shear_9km = ""
    # Check if missing values ("M" exist also:
    if "M" not in indices_dict["Lat"]:
        lat = indices_dict["Lat"]
    if "M" not in indices_dict["Lon"]:
        lon = indices_dict["Lon"]
    if "M" not in str(indices_dict["0-1 km SRH"][0]):
        srh_1km = indices_dict["0-1 km SRH"][0]
        srh_1km_abs = abs(srh_1km)
    if "M" not in str(indices_dict["0-3 km SRH"][0]):
        srh_3km = indices_dict["0-3 km SRH"][0]  
        srh_3km_abs = abs(srh_3km)
    if "M" not in str(indices_dict["SBCAPE"][0]):
        cape_sb = indices_dict["SBCAPE"][0]  
    if "M" not in str(indices_dict["MLCAPE"][0]):
        cape_ml = indices_dict["MLCAPE"][0]         
    if "M" not in str(indices_dict["MUCAPE"][0]):
        cape_mu = indices_dict["MUCAPE"][0] 
    if "M" not in str(indices_dict["SBLCL"][0]):
        lcl_sb = indices_dict["SBLCL"][0]         
    if "M" not in str(indices_dict["MLLCL"][0]):
        lcl_ml = indices_dict["MLLCL"][0]
    if "M" not in str(indices_dict["0-1 km Shear"][0]):
        shear_1km = "{:.1f}".format(indices_dict["0-1 km Shear"][0] / 1.94384)  # convert to m/s        
    if "M" not in str(indices_dict["0-3 km Shear"][0]):
        shear_3km = "{:.1f}".format(indices_dict["0-3 km Shear"][0] / 1.94384)  # convert to m/s          
    if "M" not in str(indices_dict["0-6 km Shear"][0]):
        shear_6km = "{:.1f}".format(indices_dict["0-6 km Shear"][0] / 1.94384)  # convert to m/s  
    if "M" not in str(indices_dict["0-8 km Shear"][0]):
        shear_8km = "{:.1f}".format(indices_dict["0-8 km Shear"][0] / 1.94384)  # convert to m/s  
    if "M" not in str(indices_dict["0-9 km Shear"][0]):
        shear_9km = "{:.1f}".format(indices_dict["0-9 km Shear"][0] / 1.94384)  # convert to m/s 
        
    #########################
  
    file_lines = get_list_of_problem_files(directory_in_problemfile, problem_file_name)
    problem = ""
    for ind in range(len(file_lines)):
        if inst_id in file_lines[ind] and date in file_lines[ind] and t in file_lines[ind]:
            problem = file_lines[ind].split(": ", 1)[1]
    
    #########################    
        
    # Get data in the right format for writing to excel:                
    header_values = ["0-1 km", "0-3 km", "ABS 0-1", "ABS 0-3", "SB", "ML", "MU", "SB", "ML", "0-1 km", "0-3 km", "0-6 km", "0-8 km", "0-9 km", "IOP", "Date", "Time", "Site", "Lat", "Lon", "Sounding Issues"]
    data_values = [srh_1km, srh_3km, srh_1km_abs, srh_3km_abs, cape_sb, cape_ml, cape_mu, lcl_sb, lcl_ml, shear_1km, shear_3km, shear_6km, shear_8km, shear_9km, iop, date, time, name, lat, lon, problem]

    return header_values, data_values

#########################

def write_data_to_xlsx_worksheet(workbook, worksheet, data_values, row_number):

    c = 1
    for data in data_values:
        worksheet.cell(row=row_number, column=c).value = data
        c += 1
        
    return workbook
    
#########################

def format_xlsx_worksheet(workbook, worksheet, header_values, number_of_rows):

    # Header 1
    worksheet.merge_cells('A1:D1')
    worksheet['A1'].value = "SHARPpy SRH"
    worksheet.merge_cells('E1:G1')
    worksheet['E1'].value = "SHARPpy CAPE"
    worksheet.merge_cells('H1:I1')
    worksheet['H1'].value = "SHARPpy LCL"
    worksheet.merge_cells('J1:N1')
    worksheet['J1'].value = "SHARPpy SHEAR"
    
    header1 = NamedStyle(name="header1")
    header1.font = Font(name="Calibri", size=14, bold=True)
    header1.alignment = Alignment(horizontal="center")
    
    for cell in worksheet["1:1"]:
        cell.style = header1   

    # Header 2
    header2 = NamedStyle(name="header2")
    header2.font = Font(name="Calibri", size=14, bold=True)
    header2.border = Border(bottom=Side(border_style="thick"))
    header2.alignment = Alignment(horizontal="center")

    c = 1
    for header_name in header_values:
        worksheet.cell(row=2, column=c).value = header_name
        c += 1
     
    for cell in worksheet["2:2"]:
        cell.style = header2
    
    # Data
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name="Calibri", size=14, bold=False)
    data_style.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(3,number_of_rows):
        for cell in row:
            cell.style = data_style
        
    return workbook

###############################################################################

files_to_process = get_files_from_directory(directory_in)
number_of_rows = len(files_to_process) + 2
wb = Workbook()
ws = wb.active
row_number = 3

for file in files_to_process:
    header_values, data_values = parse_info_from_indices_file(file)
    write_data_to_xlsx_worksheet(wb, ws, data_values, row_number)
    row_number += 1

format_xlsx_worksheet(wb, ws, header_values, number_of_rows)
wb.save(os.path.join(directory_out, file_out))

###############################################################################